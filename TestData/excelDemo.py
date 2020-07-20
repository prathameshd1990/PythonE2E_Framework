import openpyxl

book = openpyxl.load_workbook(
    "C:\\Users\\Saranya\\PycharmProjects\\Practice\\E2EFrameworkBuilding\\TestData\\e2eDataDriven"
    ".xlsx")
Dict = {}

sheet = book.active

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "t1":
        for j in range(2, sheet.max_column + 1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)
