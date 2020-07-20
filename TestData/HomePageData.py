import openpyxl


class HomepageData:
    test_HomePage_Data = [{"name": "Rahul", "email": "rahul@gmail.com", "password": "test@123", "gender": "Male"},
                          {"name": "Anisha", "email": "anu@gmail.com", "password": "test#123", "gender": "Female"},
                          {"name": "Raja", "email": "raja@gmail.com", "password": "test$123", "gender": "Male"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook(
            "C:\\Users\\Saranya\\PycharmProjects\\Practice\\E2EFrameworkBuilding\\TestData\\e2eDataDriven"
            ".xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == "testcaseName":
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]